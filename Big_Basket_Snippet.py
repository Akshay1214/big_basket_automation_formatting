from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

driver = webdriver.Chrome()
driver.get("https://www.bigbasket.com/pc/beauty-hygiene/health-medicine/ayurveda")
driver.maximize_window()

scroll_pause = 5
last_ht = driver.execute_script("return document.body.scrollHeight")
time.sleep(scroll_pause)
while True:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(scroll_pause)
    new_ht = driver.execute_script("return document.body.scrollHeight")
    if new_ht == last_ht:
        break
    last_ht = new_ht


result = []
products = driver.find_elements(By.CLASS_NAME, "prod-deck")
for product in products:
    pname = product.find_element(By.CSS_SELECTOR, ".prod-name > a").text
    delivery = product.find_element(By.CLASS_NAME, "delivery-opt").text
    try:
        discount_price = product.find_element(By.CLASS_NAME, "discnt-price").text
    except:
        discount_price = "No Data Found"
    try:
        without_discount = product.find_element(By.CLASS_NAME, "mp-price").text
    except:
        without_discount = "No Data Found"
    try:
        rating = product.find_element(By.CSS_SELECTOR, ".prod-name > div").text
    except:
        rating = "No Data Found"
    try:
        discount_per = product.find_element(By.CLASS_NAME, "save-price").text
    except:
        discount_per = "No Data Found"

    result.append([pname, delivery, discount_price, without_discount, rating, discount_per])

driver.quit()

wb = openpyxl.Workbook()
sheet = wb.active
for row in result:
    sheet.append(row)
wb.save("Big_Basket_Data.xlsx")

filename = "Big_Basket_Data.xlsx"
wb = load_workbook("Big_Basket_Data.xlsx")
ws1 = wb["Sheet"]

ws1["A1"] = 'Product Name'
ws1["B1"] = 'Delivery'
ws1["C1"] = 'Discount Price' 
ws1["D1"] = 'Original Price'
ws1["E1"] = 'Original Price'
ws1["F1"] = 'Discount Percent'

thin_border = Border(
                    left=Side(border_style='medium', color='FF000000'),
                    right=Side(border_style='medium', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    )

thick_border = Border(
                    left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='medium', color='FF000000'),
                    )

Double_border = Border(
                    left=Side(border_style='dashed', color='FF000000'),
                    right=Side(border_style='dashed', color='FF000000'),
                    top=Side(border_style='double', color='FF000000'),
                    bottom=Side(border_style='double', color='FF000000'),
                    )

fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00', end_color='00FFFF00')

row_num = 217
col_num = 6

row_loc = 0
col_loc = 0

for i in range(row_loc,row_loc+row_num):
    for j in range(col_loc,col_num+col_loc):
        ws1.cell(row=i+1, column=j+1).border=thin_border
        if i==row_loc:
            ws1.cell(row=i+1, column=j+1).border=Double_border
            ws1.cell(row=i+1, column=j+1).fill=fill_cell
        if i==row_loc+row_num-1:
            ws1.cell(row=i+1, column=j+1).border=thick_border

red_font = Font(color='00FF0000', bold=True, italic=True)
for cell in ws1["1:1"]:
    cell.font = red_font

h_text = Font(color = '9C0006')
h_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=h_text, fill=h_fill)
rule = Rule(type="containsText", operator="containsText", text="No Data Found", dxf=dxf)
rule.formula = ['NOT(ISERROR(SERACH("highlight", A1)))'] 
ws1.conditional_formatting.add('A1:F216', rule)

ws1.title = "Big_Basket_Data"
ws1.sheet_properties.tabcolor = '000000FF'

wb.save('Formatted_Data.xlsx')

        