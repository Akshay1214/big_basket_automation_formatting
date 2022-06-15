from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

driver = webdriver.Chrome()
driver.get("https://www.bigbasket.com/pc/beauty-hygiene/health-medicine/ayurveda")
driver.maximize_window()
time.sleep(5)
# scroll_pause = 5
# last_ht = driver.execute_script("return document.body.scrollHeight")
# time.sleep(scroll_pause)
# while True:
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#     time.sleep(scroll_pause)
#     new_ht = driver.execute_script("return document.body.scrollHeight")
#     if new_ht == last_ht:
#         break
#     last_ht = new_ht


result = []
products = driver.find_elements(By.CLASS_NAME, "prod-deck")
for product in products:
    pname = product.find_element(By.CSS_SELECTOR, ".prod-name > a").text
    delivery = product.find_element(By.CLASS_NAME, "delivery-opt").text
    try:
        discount_price = product.find_element(
            By.CLASS_NAME, "discnt-price").text
    except:
        discount_price = "No Data Found"
    try:
        without_discount = product.find_element(By.CLASS_NAME, "mp-price").text
    except:
        without_discount = "No Data Found"
    try:
        rating = product.find_element(By.CSS_SELECTOR, ".prod-name > div").text
    except:
        rating = "No Data Found"
    try:
        discount_per = product.find_element(By.CLASS_NAME, "save-price").text
    except:
        discount_per = "No Data Found"

    result.append([pname, delivery, discount_price,
                  without_discount, rating, discount_per])

driver.quit()
cell_border = Border(
    left=Side(border_style='medium', color='FF000000'),
    right=Side(border_style='medium', color='FF000000'),
    top=Side(border_style='medium', color='FF000000'),
    bottom=Side(border_style='medium', color='FF000000'),
)

header_rule = PatternFill(
    fill_type=fills.FILL_SOLID,
    start_color='00FFFF00', 
    end_color='00FFFF00'
)

highlight_rule = PatternFill(
    fill_type=fills.FILL_SOLID,
    start_color='00FF0000', 
    end_color='00FF0000'
)

wb = openpyxl.Workbook()
sh = wb.active

headers = ["Product Name", "Devlivery Date", "Discount Price", "Original Price", "Rating", "Discount Percent"]

for index, header in enumerate(headers, start=1):
    sh.cell(1, index, header)
    sh.cell(1, index).border = cell_border
    sh.cell(1, index).fill = header_rule

for row_no, row in enumerate(result, start=2):
    for col_no, val in enumerate(row, start=1):
        sh.cell(row_no, col_no, val)
        sh.cell(row_no, col_no).border = cell_border
        if val == "No Data Found" :
            sh.cell(row_no, col_no).fill = highlight_rule
wb.save("output.xlsx")